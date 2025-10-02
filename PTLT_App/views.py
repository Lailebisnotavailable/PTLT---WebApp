from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.db.models import Q
from django.views.decorators.http import require_http_methods   
import json
from django.contrib.auth.models import User
from django.db import transaction
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth import logout
from django.utils.safestring import mark_safe
from django.core.serializers.json import DjangoJSONEncoder
from datetime import time, timedelta
from django.views.decorators.http import require_POST
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.hashers import check_password
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
import datetime
import csv
import io
import traceback
from django.core.mail import send_mail
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from datetime import datetime, date
from functools import wraps
import openpyxl
import re
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework import serializers
from django.contrib.auth.hashers import make_password
# Authentication endpoint for mobile
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from .serializers import MobileClassScheduleSerializer

from .models import Account
from .models import CourseSection
from .models import ClassSchedule
from .models import AttendanceRecord
from .models import Semester
from .models import AccountUploadNotification

from collections import defaultdict
from django.utils.dateparse import parse_date


from .serializers import (
    AccountSerializer, ClassScheduleSerializer, AttendanceRecordSerializer,
    MobileAccountSerializer, MobileAttendanceSerializer
)

# for docx file
from docxtpl import DocxTemplate
from io import BytesIO
import os
from django.conf import settings

# Custom authentication decorators
def admin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        try:
            account = Account.objects.get(email=request.user.email, role='Admin')
        except Account.DoesNotExist:
            messages.error(request, "Access denied: Admin privileges required.")
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

def instructor_or_admin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        try:
            account = Account.objects.get(email=request.user.email, role__in=['Instructor', 'Admin'])
        except Account.DoesNotExist:
            messages.error(request, "Access denied: Instructor or Admin role required.")
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

def instructor_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        try:
            account = Account.objects.get(email=request.user.email, role='Instructor')
        except Account.DoesNotExist:
            messages.error(request, "Access denied: Instructor role required.")
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

#para to sa schedule ni instructor
PERIODS = [
    (time(9, 30), time(10, 20), "I"),
    (time(10, 20), time(11, 10), "II"),
    (time(11, 10), time(12, 0), "III"),
    (time(12, 0), time(12, 40), "Break"),  # Lunch
    (time(12, 40), time(13, 30), "IV"),
    (time(13, 30), time(14, 20), "V"),
    (time(14, 20), time(15, 10), "VI"),
    (time(15, 10), time(16, 0), "VII"),
]

@transaction.atomic 
def login_view(request):
    
    # Check if any accounts exist
    if not Account.objects.exists():
        # Create default admin and instructor accounts
        default_accounts = [
            {
                'user_id': '000000',
                'email': 'tupcptlt@gmail.com',
                'first_name': 'Super Admin',
                'last_name': 'Account 0',
                'role': 'Admin',
                'password': 'admin',
                'sex': 'Other',
                'status': 'Active'
            },
            {
                'user_id': '000001',
                'email': 'shelwinjay.buenaventura@gsfe.tupcavite.edu.ph',
                'first_name': 'Dummy',
                'last_name': 'Account 1',
                'role': 'Instructor',
                'password': 'instructor',
                'sex': 'Other',
                'status': 'Active'
            },
            {
                'user_id': '000002',
                'email': 'marktrieste.milan@gsfe.tupcavite.edu.ph',
                'first_name': 'Dummy',
                'last_name': 'Account 2',
                'role': 'Instructor',
                'password': 'instructor',
                'sex': 'Other',
                'status': 'Active'
            },
            {
                'user_id': '000003',
                'email': 'markjoshua.salinas@gsfe.tupcavite.edu.ph',
                'first_name': 'Dummy',
                'last_name': 'Account 3',
                'role': 'Instructor',
                'password': 'instructor',
                'sex': 'Other',
                'status': 'Active'
            },
            {
                'user_id': '000004',
                'email': 'janxander.yangco@gsfe.tupcavite.edu.ph',
                'first_name': 'Dummy',
                'last_name': 'Account 4',
                'role': 'Instructor',
                'password': 'instructor',
                'sex': 'Other',
                'status': 'Active'
            }
        ]

        for acc in default_accounts:
            # Create Django built-in User
            user = User.objects.create_user(
                username=acc['user_id'],
                email=acc['email'],
                password=acc['password'],
                first_name=acc['first_name'],
                last_name=acc['last_name']
            )

            # Save also to your custom Account model (no password needed here)
            Account.objects.create(
                user_id=acc['user_id'],
                email=acc['email'],
                first_name=acc['first_name'],
                last_name=acc['last_name'],
                role=acc['role'],
                password=None,  # Let Django User handle passwords
                sex=acc['sex'],
                status=acc['status'],
                course_section=None
            )
    
    if request.method == 'POST':
        #VVVV get the email and password inputted by the user
        email = request.POST.get('email')
        password = request.POST.get('password')

        try:
            # First check if this is a temporary password login
            account = Account.objects.get(email=email)
            
            # Check if user has Django User account
            try:
                user_obj = User.objects.get(email=email)
                # Try normal authentication first
                user = authenticate(request, username=user_obj.username, password=password)
                if user is not None:
                    # Check if password is the temporary "00000" or "admin"
                    if password == "000000" or password == "admin":
                        # Store user info in session for password change
                        request.session['temp_user_id'] = account.user_id
                        request.session['temp_email'] = email
                        messages.info(request, "You must change your password before continuing.")
                        return redirect('force_password_change')
                    
                    # Normal login
                    login(request, user)
                    request.session['user_id'] = account.user_id
                    request.session['role'] = account.role

                    # Redirect based on role
                    if account.role == 'Admin':
                        return redirect('account_management')
                    elif account.role == 'Instructor':
                        return redirect('schedule')
                    else:
                        messages.error(request, "Unknown user role.")
                        return redirect('login')
                
            except User.DoesNotExist:
                # No Django User exists, but Account exists with temp password
                if password == "00000":
                    # Store account info for password setup
                    request.session['temp_user_id'] = account.user_id
                    request.session['temp_email'] = email
                    messages.info(request, "Please set up your password to continue.")
                    return redirect('force_password_change')
                else:
                    messages.error(request, "Account not fully set up. Please contact administrator.")
                    return redirect('login')
            
            messages.error(request, "Invalid credentials")
            return redirect('login')

        except Account.DoesNotExist:
            messages.error(request, "No account found with that email.")
            return redirect('login')
        
    return render(request, 'login.html')


def force_password_change(request):
    # Check if user came from login with temp password
    if 'temp_user_id' not in request.session:
        messages.error(request, "Unauthorized access.")
        return redirect('login')
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if new_password == '000000':
            messages.error(request, "Please try a different password.")
            return render(request, 'force_password_change.html')
        
        if new_password == 'secret':
            messages.error(request, "Uy bat mo alam?")
            return render(request, 'force_password_change.html')
        
        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'force_password_change.html')
        
        if len(new_password) < 6:  # Add your password requirements
            messages.error(request, "Password must be at least 6 characters long.")
            return render(request, 'force_password_change.html')
        
        try:
            user_id = request.session['temp_user_id']
            email = request.session['temp_email']
            account = Account.objects.get(user_id=user_id, email=email)
            
            # Create or update Django User
            try:
                user_obj = User.objects.get(email=email)
                # Update existing user password
                user_obj.set_password(new_password)
                user_obj.save()
            except User.DoesNotExist:
                # Create new Django User
                user_obj = User.objects.create_user(
                    username=account.user_id,
                    email=account.email,
                    password=new_password,
                    first_name=account.first_name,
                    last_name=account.last_name
                )
            
            # Clear temporary session data
            del request.session['temp_user_id']
            del request.session['temp_email']
            
            # Auto-login the user
            user = authenticate(request, username=user_obj.username, password=new_password)
            if user:
                login(request, user)
                request.session['user_id'] = account.user_id
                request.session['role'] = account.role
                
                messages.success(request, "Password updated successfully!")
                
                # Redirect based on role
                if account.role == 'Admin':
                    return redirect('account_management')
                elif account.role == 'Instructor':
                    return redirect('schedule')
                else:
                    return redirect('login')
            
        except Account.DoesNotExist:
            messages.error(request, "Account not found.")
            return redirect('login')
    
    return render(request, 'force_password_change.html')

def logout_view(request):
    logout(request)  # Destroys session and logs out user
    return redirect('login') 

@admin_required
def create_instructor(request):
    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'instructor_form':
            # Instructor form submitted
            user_id = request.POST.get('user_id')
            email = request.POST.get('email')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            role = request.POST.get('role')
            password = request.POST.get('password')
            sex = request.POST.get('sex')

            if Account.objects.filter(user_id=user_id).exists():
                messages.error(request, 'User ID already exists.')
            elif Account.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists.')
            else:
                Account.objects.create(
                    user_id=user_id,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    role=role,
                    password=password,
                    sex=sex
                )
                messages.success(request, 'Instructor account created successfully!')
                return redirect('create_instructor')

        elif form_type == 'course_section_form':
            # Course & section form submitted
            course_name = request.POST.get('course_name')
            section_name = request.POST.get('section_name')

            try:
                course_section = CourseSection.objects.create(
                    course_name=course_name,
                    section_name=section_name
                )
                messages.success(request, f'Course Section "{course_section.course_section}" created successfully!')
                return redirect('create_instructor')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')

    return render(request, 'create_instructor.html')

def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get('email')

        try:
            # Fetch the user using the default User model
            user = User.objects.get(email=email)

            # Generate the password reset token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            current_site = get_current_site(request).domain
            # Generate the reset password URL
            reset_link = f"http://{current_site}/reset-password/{uid}/{token}/"

            # HTML email content - using triple quotes without f-string for CSS, then formatting
            email_subject = 'Password Reset Request'
            
            # Create the email body using format() method instead of f-string
            email_body = """
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>Password Reset</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 0;
                        padding: 0;
                        background-color: #f5f5f5;
                    }}
                    .container {{
                        width: 100%;
                        max-width: 600px;
                        margin: 0 auto;
                        background-color: #ffffff;
                        padding: 20px;
                        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                    }}
                    h1 {{
                        color: #333;
                        text-align: center;
                    }}
                    p {{
                        font-size: 1rem;
                        line-height: 1.5;
                        color: #555;
                    }}
                    .button {{
                        display: inline-block;
                        padding: 12px 24px;
                        background-color: #661e1e;
                        color: white !important;
                        text-decoration: none;
                        border-radius: 4px;
                        font-size: 1rem;
                        margin-top: 20px;
                        text-align: center;
                        transition: background-color 0.3s ease;
                    }}
                    .button:hover {{
                        background-color: #a74545;
                    }}
                    .footer {{
                        margin-top: 30px;
                        padding-top: 20px;
                        border-top: 1px solid #eee;
                        font-size: 0.9rem;
                        color: #777;
                    }}
                    .warning {{
                        background-color: #fff3cd;
                        border: 1px solid #ffeaa7;
                        border-radius: 4px;
                        padding: 15px;
                        margin: 20px 0;
                        color: #856404;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header-accent"></div>
                    <h1>Password Reset Request</h1>
                    <p>Hello {first_name},</p>
                    <p>We received a request to reset your password for your account. To proceed with resetting your password, please click the button below:</p>
                    <p><a href="{reset_link}" class="button">Reset Password</a></p>
                    
                    <div class="warning">
                        <strong>Security Notice:</strong> This link will expire in 24 hours for your security. If you didn't request this password reset, please ignore this email and your password will remain unchanged.
                    </div>
                    
                    <div class="footer">
                        <p>If the button above doesn't work, you can copy and paste this link into your browser:</p>
                        <p style="word-break: break-all; color: #dc2626 !important; font-weight: 600; text-decoration: none !important;">{reset_link}</p>
                        <p style="margin-top: 20px;">
                            <span style="color: #6b7280;">Best regards,</span><br>
                            <strong style="color: #dc2626;">PTLT TUP-CAVITE</strong>
                        </p>
                    </div>
                </div>
            </body>
            </html>
            """.format(first_name=user.first_name, reset_link=reset_link)

            # Send the HTML email
            send_mail(
                email_subject,
                '',  # Plain text version of the email (empty since we are sending HTML)
                'from@example.com',  # Set your sender email
                [email],
                fail_silently=False,
                html_message=email_body  # HTML version of the email
            )

            # Show success message to the user
            messages.success(request, 'Password reset link has been sent to your email address.')
            return redirect('login')

        except User.DoesNotExist:
            # Handle the case where the user doesn't exist
            messages.error(request, 'No account found with this email address.')
    return render(request, 'forgot_password.html')

def reset_password(request, encoded_email, token):
    try:
        # Decode the user ID from the encoded email
        try:
            uid = urlsafe_base64_decode(encoded_email).decode('utf-8')
            print(f"Decoded user ID: {uid}")
        except Exception as e:
            print(f"Error decoding email: {e}")
            messages.error(request, 'Invalid or expired reset link.')
            return redirect('login')
        
        # Fetch the user based on the decoded ID
        try:
            user = User.objects.get(pk=uid)
        except User.DoesNotExist:
            print(f"User does not exist for ID: {uid}")
            messages.error(request, 'User not found.')
            return redirect('login')
        
        # Check if the token matches the user's reset token
        if default_token_generator.check_token(user, token):
            if request.method == 'POST':
                new_password = request.POST.get('password')
                confirm_password = request.POST.get('confirm_password')
                print(f"new_password: {new_password}, confirm_password: {confirm_password}")

                # Ensure the passwords match
                if new_password == confirm_password:
                    user.set_password(new_password)  # Set the new password
                    user.save()  # Save the user object
                    print("Password reset successfully!")
                    messages.success(request, 'Your password has been reset successfully!')
                    update_session_auth_hash(request, user)  # Keep the user logged in
                    return redirect('login')  # Redirect to login page
                else:
                    messages.error(request, 'Passwords do not match. Please try again.')

            return render(request, 'reset_password.html', {'uid': encoded_email, 'token': token})

        else:
            print(f"Invalid token for user: {uid}")
            messages.error(request, 'Invalid or expired reset link.')
            return redirect('login')

    except Exception as e:
        print(f"Error occurred: {e}")
        messages.error(request, 'An error occurred during password reset. Please try again later.')
        return redirect('login')
    

@instructor_required
def student_attendance_records(request):
    # Get logged-in instructor's Account entry
    try:
        instructor_account = Account.objects.get(email=request.user.email, role='Instructor')
    except Account.DoesNotExist:
        return render(request, "error.html", {"message": "Instructor account not found"})

    # Subjects/Courses taught by this instructor
    schedules = ClassSchedule.objects.filter(professor=instructor_account)

    selected_schedule_id = request.GET.get("schedule")
    selected_date_range = request.GET.get("date_range")
    date_ranges = []
    attendance_table = []

    if selected_schedule_id:
        # Fetch unique attendance dates for selected class schedule
        attendance_dates = AttendanceRecord.objects.filter(
            class_schedule_id=selected_schedule_id
        ).values_list("date", flat=True).distinct().order_by("date")

        attendance_dates = list(attendance_dates)

        # Group into 8-day ranges for the filter dropdown
        for i in range(0, len(attendance_dates), 8):
            start_date = attendance_dates[i]
            end_date = attendance_dates[min(i + 7, len(attendance_dates) - 1)]
            date_ranges.append({
                "value": f"{start_date}_to_{end_date}",
                "label": f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
            })

        if selected_date_range:
            try:
                start_str, end_str = selected_date_range.split("_to_")
                start_date = parse_date(start_str)
                end_date = parse_date(end_str)
            except (ValueError, TypeError):
                start_date = end_date = None

            if start_date and end_date:
                # Get all attendance records within the date range
                attendance_qs = AttendanceRecord.objects.filter(
                    class_schedule_id=selected_schedule_id,
                    date__range=(start_date, end_date)
                ).select_related('student')

                # Get schedule object once
                schedule_obj = ClassSchedule.objects.get(id=selected_schedule_id)

                # Get students in the same course_section as the schedule
                students_in_schedule = Account.objects.filter(
                    course_section_id=schedule_obj.course_section_id
                ).order_by('last_name', 'first_name')

                # Map attendance data: {student_id: {date: {'status': status, 'time_in': time_in, 'time_out': time_out}}}
                attendance_data = defaultdict(lambda: defaultdict(dict))
                for record in attendance_qs:
                    attendance_data[record.student.id][record.date] = {
                        'status': record.status,
                        'time_in': record.time_in,
                        'time_out': record.time_out
                    }

                # Build date headers (max 8 dates)
                date_headers = [d for d in attendance_dates if start_date <= d <= end_date][:8]
                num_empty_date_column = 8 - len(date_headers)
                #create a pseudo list just for the for loop in html to work
                num_empty_date_columns = []
                for i in range(num_empty_date_column):
                    num_empty_date_columns.append("")


                attendance_table = []
                for student in students_in_schedule:
                    course_section_for_student = student.course_section
                    # Build dates_statuses as a list of dicts containing all attendance info
                    dates_statuses = []
                    for date in date_headers:
                        attendance_info = attendance_data[student.id].get(date, {})
                        dates_statuses.append({
                            'status': attendance_info.get('status', ''),
                            'time_in': attendance_info.get('time_in', ''),
                            'time_out': attendance_info.get('time_out', '')
                        })
                    
                    # Pad with empty dicts to always have 8 items
                    while len(dates_statuses) < 8:
                        dates_statuses.append({
                            'status': '',
                            'time_in': '',
                            'time_out': ''
                        })

                    # DEBUG: Check the length
                    print(f"Student {student.user_id}: dates_statuses length = {len(dates_statuses)}")

                    row = {
                        "student_id": student.user_id, 
                        "name": f"{student.first_name} {student.last_name}",
                        "sex": student.sex,
                        "course": course_section_for_student,
                        "subject": schedule_obj.course_code,
                        "room": schedule_obj.room_assignment,
                        "dates": dates_statuses,  # Always 8 items now
                    }
                    attendance_table.append(row)

                context = {
                    "schedules": schedules,
                    "date_ranges": date_ranges,
                    "selected_schedule_id": selected_schedule_id,
                    "selected_date_range": selected_date_range,
                    "attendance_table": attendance_table,
                    "date_headers": date_headers,
                    "num_empty_date_columns": num_empty_date_columns,
                }
                return render(request, "student_attendance_records.html", context)

        # If no valid date range selected or dates invalid, still render with schedules & date ranges
        context = {
            "schedules": schedules,
            "date_ranges": date_ranges,
            "selected_schedule_id": selected_schedule_id,
            "attendance_table": [],
            "date_headers": [],
            "num_empty_date_columns": ["", "", "", "", "", "", "", "", ],
        }
        return render(request, "student_attendance_records.html", context)

    # If no schedule selected at all, render with just schedules and empty date ranges
    context = {
        "schedules": schedules,
        "date_ranges": date_ranges,
        "selected_schedule_id": selected_schedule_id,
        "attendance_table": [],
        "date_headers": [],
        "num_empty_date_columns": ["", "", "", "", "", "", "", "", ],
    }
    return render(request, "student_attendance_records.html", context)



@require_POST
@instructor_required
def update_class_schedule_instructor(request):
    try:
        data = json.loads(request.body)

        # Validate required fields
        required_fields = ['course_code', 'time_in', 'time_out', 'room_assignment', 'grace_period']
        for field in required_fields:
            if field not in data:
                return JsonResponse({'success': False, 'error': f'Missing field: {field}'}, status=400)

        # Match instructor via user_id
        instructor = Account.objects.get(user_id=request.user.username, role='Instructor')

        # Get the specific schedule (only one per instructor + course_code assumed)
        schedule = ClassSchedule.objects.get(course_code=data['course_code'], professor=instructor)

        # Apply updates
        schedule.time_in = data['time_in']
        schedule.time_out = data['time_out']
        schedule.room_assignment = data['room_assignment']
        schedule.grace_period = int(data['grace_period'])
        schedule.save()

        return JsonResponse({'success': True})

    except ClassSchedule.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Class schedule not found.'}, status=404)

    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Instructor account not found.'}, status=403)

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data.'}, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@instructor_required    
def instructor_schedule(request):
    user = request.user
    try:
        instructor = Account.objects.get(user_id=user.username, role='Instructor')
    except Account.DoesNotExist:
        return render(request, 'error.html', {'message': 'Instructor not found.'})

    schedules = ClassSchedule.objects.filter(professor=instructor)

    for schedule in schedules:
        student_acc = Account.objects.filter(course_section_id=schedule.course_section_id)
        student_count = len(student_acc)
        schedule.student_count = int(student_count)
        schedule.save()

    return render(request, 'schedule.html', {
        'class_schedules': schedules,
    })

@admin_required
def account_management(request):
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')

    accounts = Account.objects.all()

    if role_filter:
        accounts = accounts.filter(role__iexact=role_filter)
    if status_filter:
        accounts = accounts.filter(status__iexact=status_filter)
    if search_query:
        accounts = accounts.filter(
            Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query)
        )

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('partials/account_table_body.html', {'accounts': accounts})
        return JsonResponse({'html': html})

    return render(request, 'account_management.html', {'accounts': accounts})

@csrf_exempt
@admin_required
def delete_account(request, account_id):
    if request.method == 'POST':
        acc = get_object_or_404(Account, id=account_id)
        acc.delete()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@csrf_exempt
@admin_required
def update_account(request, account_id):
    print("1")
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print("Received data:", data)  # 👈 Debug
            
            account = get_object_or_404(Account, id=account_id)

            # Direct assignment from data
            account.user_id = data.get('user_id', account.user_id)
            account.first_name = data.get('first_name', account.first_name)
            account.last_name = data.get('last_name', account.last_name)
            account.role = data.get('role', account.role)
            account.email = data.get('email', account.email)

            account.save()

            return JsonResponse({'status': 'success'})
        except Exception as e:
            print("Error updating account:", str(e))
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'invalid_request'})

import csv
import io
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Account, ClassSchedule, CourseSection

@csrf_exempt
@admin_required
def import_class_schedule(request):
    print("⚡ Import request received")
    print("Method:", request.method)
    print("FILES:", request.FILES)
    print("POST:", request.POST)

    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Only POST method allowed."}, status=400)

    if "csv_file" not in request.FILES:
        return JsonResponse({"status": "error", "message": "No CSV file uploaded."}, status=400)

    csv_file = request.FILES["csv_file"]
    print(f"📂 Received file: {csv_file.name}, size={csv_file.size} bytes")

    try:
        data = csv_file.read().decode("utf-8")
        io_string = io.StringIO(data)
        reader = csv.DictReader(io_string)
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Failed to read CSV: {e}"}, status=400)

    results = {
        "imported": 0,
        "skipped": 0,
        "errors": []
    }

    line_num = 1  # header = line 1
    for row in reader:
        line_num += 1
        print(f"[IMPORT] Processing line {line_num}: {row}")

        prof_user_id = row.get("professor_user_id")
        professor = None
        if prof_user_id:
            professor = Account.objects.filter(user_id=prof_user_id, role="Instructor").first()

        if not professor:
            msg = f"Line {line_num}: Professor with user_id '{prof_user_id}' not found in accounts."
            print("❌", msg)
            results["skipped"] += 1
            results["errors"].append(msg)
            continue

        section_id = row.get("course_section_id")
        course_section = None
        if section_id:
            course_section = CourseSection.objects.filter(id=section_id).first()

        if not course_section:
            msg = f"Line {line_num}: Section '{section_id}' not found in CourseSection."
            print("❌", msg)
            results["skipped"] += 1
            results["errors"].append(msg)
            continue

        try:
            ClassSchedule.objects.create(
                professor=professor,
                course_title=row.get("course_title"),
                course_code=row.get("course_code"),
                course_section=course_section,
                time_in=row.get("time_in"),
                time_out=row.get("time_out"),
                days=row.get("days"),
                grace_period=row.get("grace_period") or 0,
                student_count=row.get("student_count") or 0,
                remote_device=row.get("remote_device"),
                room_assignment=row.get("room_assignment"),
            )
            results["imported"] += 1
            print(f"✅ Line {line_num}: Schedule imported successfully.")
        except Exception as e:
            msg = f"Line {line_num}: Failed to save schedule. Error: {str(e)}"
            print("❌", msg)
            results["skipped"] += 1
            results["errors"].append(msg)

    if results["imported"] == 0:
        status = "failed"  # Nothing imported
    elif results["skipped"] > 0:
        status = "partial"  # Some succeeded, some failed
    else:
        status = "ok"  # Everything succeeded
    return JsonResponse({
        "status": status,
        "imported": results["imported"],
        "skipped": results["skipped"],
        "errors": results["errors"],
    })
#try pdf import
@require_http_methods(["POST"])
def import_class_from_excel(request):
    if 'excel_file' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'No Excel file provided'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Initialize data containers
        schedule_data = {}
        all_students = []
        
        # Process each sheet to find schedule and student information
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Check if this sheet contains schedule information
            first_cell = sheet.cell(1, 1).value
            if first_cell and 'Schedule ID' in str(first_cell):
                # Parse the schedule information from the merged cell
                text = str(first_cell)
                
                # Extract Schedule ID
                schedule_id_match = re.search(r'Schedule ID\s*:\s*([A-Z0-9]+)', text)
                if schedule_id_match and 'schedule_id' not in schedule_data:
                    schedule_data['schedule_id'] = schedule_id_match.group(1).strip()
                
                # Extract Subject (Course Code - Course Title)
                subject_match = re.search(r'Subject\s*:\s*([A-Z0-9-]+)\s*-\s*(.+?)(?:\s{2,}|Venue)', text)
                if subject_match and 'course_code' not in schedule_data:
                    schedule_data['course_code'] = subject_match.group(1).strip()
                    schedule_data['course_title'] = subject_match.group(2).strip()
                
                # Extract Day/Time
                day_time_match = re.search(r'Day/Time\s*:\s*([MTWRFSU])\s+(\d{1,2}:\d{2}[AP]M)-(\d{1,2}:\d{2}[AP]M)', text)
                if day_time_match and 'day' not in schedule_data:
                    day_map = {
                        'M': 'Monday', 'T': 'Tuesday', 'W': 'Wednesday',
                        'R': 'Thursday', 'F': 'Friday', 'S': 'Saturday', 'U': 'Sunday'
                    }
                    schedule_data['day'] = day_map.get(day_time_match.group(1), 'Monday')
                    
                    time_in_str = day_time_match.group(2)
                    time_out_str = day_time_match.group(3)
                    
                    schedule_data['time_in'] = datetime.strptime(time_in_str, '%I:%M%p').time()
                    schedule_data['time_out'] = datetime.strptime(time_out_str, '%I:%M%p').time()
                
                # Extract Course/Section
                section_match = re.search(r'Course/Section\s*:\s*(.+?)(?:\s{2,}|$)', text)
                if section_match and 'section_name' not in schedule_data:
                    section_str = section_match.group(1).strip()
                    # Handle format like "BET-COET-C-BET-COET-C-4A-C"
                    parts = section_str.split('-')
                    if len(parts) >= 3:
                        # Extract course and section
                        # For "BET-COET-C-BET-COET-C-4A-C", we want "BET-COET-C" and "4A-C"
                        mid_point = len(parts) // 2
                        schedule_data['course_name'] = '-'.join(parts[:mid_point])
                        schedule_data['section_name'] = '-'.join(parts[-2:])
            
            # Check if this sheet contains student information
            elif sheet.max_row > 1:
                # Check if first row has student headers
                row1_values = [str(sheet.cell(1, col).value or '') for col in range(1, min(6, sheet.max_column + 1))]
                if 'Student No.' in ' '.join(row1_values):
                    # This is a student list sheet
                    for row_idx in range(2, sheet.max_row + 1):
                        # Get student data from columns
                        student_no_cell = sheet.cell(row_idx, 2).value  # Column B
                        name_cell = sheet.cell(row_idx, 3).value  # Column C
                        
                        if not student_no_cell or not name_cell:
                            continue
                        
                        # Skip if it's a total row
                        if 'Total' in str(student_no_cell):
                            break
                        
                        # Extract ID (TUPC-22-0352 → 220352)
                        student_no = str(student_no_cell).strip()
                        match = re.match(r'TUPC-(\d{2})-(\d{4})', student_no)
                        if not match:
                            continue
                        
                        short_id = match.group(1) + match.group(2)  # e.g., "220352"
                        
                        # Parse name (LASTNAME, FIRSTNAME MIDDLENAME)
                        name = str(name_cell).strip()
                        name_parts = name.split(',')
                        
                        if len(name_parts) >= 2:
                            last_name = name_parts[0].strip().title()
                            first_name_parts = name_parts[1].strip().split()
                            first_name = first_name_parts[0].title() if first_name_parts else ''
                            
                            all_students.append({
                                'user_id': short_id,
                                'first_name': first_name,
                                'last_name': last_name
                            })
        
        # Validate we got the required schedule data
        required_fields = ['course_code', 'course_title', 'day', 'time_in', 'time_out', 'course_name', 'section_name']
        missing_fields = [f for f in required_fields if f not in schedule_data]
        
        if missing_fields:
            return JsonResponse({
                'status': 'error',
                'message': f'Could not parse schedule information. Missing: {", ".join(missing_fields)}'
            }, status=400)
        
        if not all_students:
            return JsonResponse({
                'status': 'error',
                'message': 'Could not find any students in the Excel file'
            }, status=400)
        
        # Create or get CourseSection
        course_section, created = CourseSection.objects.get_or_create(
            course_name=schedule_data['course_name'],
            section_name=schedule_data['section_name']
        )
        
        # Create ClassSchedule
        class_schedule = ClassSchedule.objects.create(
            course_code=schedule_data['course_code'],
            course_title=schedule_data['course_title'],
            time_in=schedule_data['time_in'],
            time_out=schedule_data['time_out'],
            days=schedule_data['day'],
            course_section=course_section,
            professor=None,
            student_count=0,
            grace_period=15,
            remote_device='',
            room_assignment='-'
        )
        
        # Create student accounts
        created_students = 0
        skipped_students = 0
        
        for student_info in all_students:
            if not Account.objects.filter(user_id=student_info['user_id']).exists():
                Account.objects.create(
                    user_id=student_info['user_id'],
                    email='',  # Empty - will be filled via mobile app
                    first_name=student_info['first_name'],
                    last_name=student_info['last_name'],
                    role='Student',
                    password='00000',
                    sex='',
                    status='Pending',
                    course_section=course_section,
                    fingerprint_template=None
                )
                created_students += 1
            else:
                skipped_students += 1
        
        # Update student count
        class_schedule.student_count = len(all_students)
        class_schedule.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Successfully imported class schedule',
            'details': {
                'course_code': schedule_data['course_code'],
                'course_title': schedule_data['course_title'],
                'course_section': course_section.course_section,
                'day': schedule_data['day'],
                'time': f"{schedule_data['time_in']} - {schedule_data['time_out']}",
                'students_created': created_students,
                'students_skipped': skipped_students,
                'total_students': len(all_students)
            }
        })
        
    except Exception as e:
        print(f"Error importing Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': f'Failed to parse Excel: {str(e)}'
        }, status=500)

@admin_required
def class_management(request):
    
    today = timezone.now().date()
    
    current_semester = Semester.objects.filter(
        start_date__lte=today,
        end_date__gte=today
    ).first()
    active_semester = current_semester
    
    # Get unread notifications count
    new_accounts_count = AccountUploadNotification.objects.filter(is_read=False).count()
    recent_uploads = AccountUploadNotification.objects.filter(is_read=False)[:5]  # Last 5
    
    #update student count

    schedules = ClassSchedule.objects.all()
    for schedule in schedules:
        student_acc = Account.objects.filter(course_section_id=schedule.course_section_id)
        student_count = len(student_acc)
        schedule.student_count = int(student_count)
        schedule.save()

    # Mark as read if user clicks "Mark as Read"
    if request.GET.get('mark_read') == 'true':
        AccountUploadNotification.objects.filter(is_read=False).update(is_read=True)
        messages.success(request, f'Marked {new_accounts_count} notifications as read.')
        return redirect('class_management')
    
    course_sections = CourseSection.objects.all()

    if request.method == 'POST':
        course_code = request.POST.get('course_code')
        course_name = request.POST.get('course_name')
        time_in = request.POST.get('time_in')
        time_out = request.POST.get('time_out')
        day = request.POST.get('day')
        course_section_str = request.POST.get('course_section')
        remote_device = request.POST.get('remote_device')

        try:
            section_obj = CourseSection.objects.get(course_section=course_section_str)
        except CourseSection.DoesNotExist:
            section_obj = None

        ClassSchedule.objects.create(
            course_code=course_code,
            course_title=course_name,
            time_in=time_in,
            time_out=time_out,
            days=day,
            course_section=section_obj,
            professor=None,
            student_count=0,
            grace_period=0,
            remote_device=remote_device,
            room_assignment='-',
        )

    classes = ClassSchedule.objects.all()

    # Get instructors only
    instructors = Account.objects.filter(role="Instructor").values("first_name", "last_name")
    instructors_json = json.dumps(list(instructors), cls=DjangoJSONEncoder)

    return render(request, 'class_management.html', {
        "active_semester": active_semester,
        'course_sections': course_sections,
        'classes': classes,
        'instructors_json': instructors_json,
        'new_accounts_count': new_accounts_count,  # NEW
        'recent_uploads': recent_uploads,  # NEW
        'current_semester': current_semester
    })

@require_http_methods(["POST"])
@admin_required
def add_course_section(request):
    try:
        data = json.loads(request.body)
        course_name = data.get('course_name', '').strip()
        section_name = data.get('section_name', '').strip()

        if not course_name or not section_name:
            return JsonResponse({
                'status': 'error',
                'message': 'Course name and section name are required.'
            }, status=400)

        # Check if already exists
        course_section_str = f"{course_name} {section_name}"
        if CourseSection.objects.filter(course_section=course_section_str).exists():
            return JsonResponse({
                'status': 'error',
                'message': f'Section "{course_section_str}" already exists.'
            }, status=400)

        # Create new course section
        new_section = CourseSection.objects.create(
            course_name=course_name,
            section_name=section_name
        )

        return JsonResponse({
            'status': 'success',
            'message': 'Course section added successfully.',
            'course_section': new_section.course_section
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)    

@csrf_exempt
@admin_required
def update_class_schedule(request, pk):
    if request.method == "POST":
        try:
            cls = ClassSchedule.objects.get(id=pk)
            data = json.loads(request.body)

            prof_name = data.get("professor_name", "").strip()
            if prof_name:
                try:
                    first, last = prof_name.split(" ", 1)
                    professor = Account.objects.get(first_name=first, last_name=last)
                    cls.professor = professor
                except Account.DoesNotExist:
                    cls.professor = None

            cls.time_in = data.get("time_in")
            cls.time_out = data.get("time_out")
            cls.days = data.get("day")
            cls.save()

            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

@csrf_exempt
@admin_required
def delete_class_schedule(request, pk):
    if request.method == "POST":
        try:
            cls = ClassSchedule.objects.get(id=pk)
            cls.delete()
            return JsonResponse({"status": "deleted"})
        except:
            return JsonResponse({"status": "error"}, status=400)

@instructor_or_admin_required
def attendance_report_template(request):
    # Get logged-in instructor's Account entry
    try:
        instructor_account = Account.objects.get(email=request.user.email, role='Instructor')
    except Account.DoesNotExist:
        return render(request, "error.html", {"message": "Instructor account not found"})
    
    # Get class schedules for the dropdown
    schedules = ClassSchedule.objects.filter(professor=instructor_account)
    
    selected_schedule_id = request.GET.get("schedule")
    selected_date_range = request.GET.get("date_range")
    date_ranges = []
    attendance_table = []
    attendance_data = {}
    students_list = []
    
    if selected_schedule_id:
        try:
            # Get the selected schedule
            schedule_obj = ClassSchedule.objects.get(id=selected_schedule_id)
            
            # Class details for the form
            attendance_data_attendance_report = {
                'subject': schedule_obj.course_title,
                'faculty_name': f"{schedule_obj.professor.first_name} {schedule_obj.professor.last_name}" if schedule_obj.professor else 'TBA',
                'course': schedule_obj.course_section.course_name if schedule_obj.course_section else '',
                'room': schedule_obj.room_assignment or 'TBA',
                'year_section': schedule_obj.course_section.section_name if schedule_obj.course_section else '',
                'schedule': f"{schedule_obj.days} {schedule_obj.time_in}-{schedule_obj.time_out}",
            }
            
        except ClassSchedule.DoesNotExist:
            pass

        # Fetch unique attendance dates for selected class schedule
        attendance_dates = AttendanceRecord.objects.filter(
            class_schedule_id=selected_schedule_id
        ).values_list("date", flat=True).distinct().order_by("date")

        attendance_dates = list(attendance_dates)

        # Group into 8-day ranges for the filter dropdown
        for i in range(0, len(attendance_dates), 8):
            start_date = attendance_dates[i]
            end_date = attendance_dates[min(i + 7, len(attendance_dates) - 1)]
            date_ranges.append({
                "value": f"{start_date}_to_{end_date}",
                "label": f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
            })

        #if may sinelect na na date range
        if selected_date_range:
            try:
                start_str, end_str = selected_date_range.split("_to_")
                start_date = parse_date(start_str)
                end_date = parse_date(end_str)
            except (ValueError, TypeError):
                start_date = end_date = None

            if start_date and end_date:
                # Get all attendance records within the date range
                attendance_qs = AttendanceRecord.objects.filter(
                    class_schedule_id=selected_schedule_id,
                    date__range=(start_date, end_date)
                ).select_related('student')

                # Get schedule object once
                schedule_obj = ClassSchedule.objects.get(id=selected_schedule_id)

                # Get students in the same course_section as the schedule
                students_in_schedule = Account.objects.filter(
                    course_section_id=schedule_obj.course_section_id
                ).order_by('last_name', 'first_name')

                # Map attendance data: {student_id: {date: {'status': status, 'time_in': time_in, 'time_out': time_out}}}
                attendance_data = defaultdict(lambda: defaultdict(dict))
                for record in attendance_qs:
                    attendance_data[record.student.id][record.date] = {
                        'status': record.status,
                        'time_in': record.time_in,
                        'time_out': record.time_out
                    }

                # Build date headers (max 8 dates)
                date_headers = [d for d in attendance_dates if start_date <= d <= end_date][:8]
                num_empty_date_column = 8 - len(date_headers)
                #create a pseudo list just for the for loop in html to work
                num_empty_date_columns = []
                for i in range(num_empty_date_column):
                    num_empty_date_columns.append("")


                attendance_table = []
                for student in students_in_schedule:
                    course_section_for_student = student.course_section
                    # Build dates_statuses as a list of dicts containing all attendance info
                    dates_statuses = []
                    for date in date_headers:
                        attendance_info = attendance_data[student.id].get(date, {})
                        dates_statuses.append({
                            'status': attendance_info.get('status', ''),
                            'time_in': attendance_info.get('time_in', ''),
                            'time_out': attendance_info.get('time_out', '')
                        })
                    
                    # Pad with empty dicts to always have 8 items
                    while len(dates_statuses) < 8:
                        dates_statuses.append({
                            'status': '',
                            'time_in': '',
                            'time_out': ''
                        })
                    
                        
                    row = {
                        "student_id": student.user_id, 
                        "name": f"{student.first_name} {student.last_name}",
                        "sex": student.sex,
                        "course": course_section_for_student,
                        "subject": schedule_obj.course_code,
                        "room": schedule_obj.room_assignment,
                        "dates": dates_statuses,  # Always 8 items now
                    }
                    attendance_table.append(row)

                else:
                    row = [""]
                # Build date headers (max 8 dates)
                date_headers = [d for d in attendance_dates if start_date <= d <= end_date][:8]
                num_empty_date_column = 8 - len(date_headers)
                #create a pseudo list just for the for loop in html to work
                num_empty_date_columns = []
                for i in range(num_empty_date_column):
                    num_empty_date_columns.append("")

                num_empty_row = 40 - len(attendance_table)
                #create a pseudo list just for the for loop in html to work
                num_empty_rows = []
                num_filled_rows = []
                for i in range(num_empty_row):
                    num_empty_rows.append("")

                for i in range(len(attendance_table)):
                    num_filled_rows.append("")

                print(f"{len(num_empty_rows)}")
                print(f"{len(num_filled_rows)}")

                context = {
                    'attendance_data': attendance_data_attendance_report,
                    "schedules": schedules,
                    "date_ranges": date_ranges,
                    "selected_schedule_id": selected_schedule_id,
                    "selected_date_range": selected_date_range,
                    "attendance_table": attendance_table,
                    "date_headers": date_headers,
                    "num_empty_rows": num_empty_rows,
                    "num_filled_rows": num_filled_rows,
                    "num_empty_date_columns": num_empty_date_columns,
                }
                return render(request, "attendance_report_template.html", context)
        else:
    # If no valid date range selected or dates invalid, still render with schedules & date ranges
            context = {
                'attendance_data': attendance_data_attendance_report,
                "schedules": schedules,
                "date_ranges": date_ranges,
                "selected_schedule_id": selected_schedule_id,
                "attendance_table": [],
                "date_headers": [],
                "num_empty_rows": [""]*40,
                "num_filled_rows": [],
                "num_empty_date_columns": ["", "", "", "", "", "", "", "", ],
            }
            return render(request, "attendance_report_template.html", context)

    # If no schedule selected at all, render with just schedules and empty date ranges
    context = {
        'attendance_data': {"":"",'faculty_name': f"{instructor_account.first_name} {instructor_account.last_name}" if instructor_account else 'TBA',"":"","":"","":"","":""},
        "schedules": schedules,
        "date_ranges": date_ranges,
        "selected_schedule_id": selected_schedule_id,
        "attendance_table": [],
        "date_headers": [],
        "num_empty_rows": [""]*40,
        "num_filled_rows": [],
        "num_empty_date_columns": ["", "", "", "", "", "", "", "", ],
    }
    return render(request, "attendance_report_template.html", context)
    

@admin_required
def set_semester(request):
    today = date.today()
    current_semester = Semester.objects.filter(start_date__lte=today, end_date__gte=today).first()

    if request.method == "POST":
        start = request.POST.get("semester_start")
        end = request.POST.get("semester_end")

        if not start or not end:
            messages.error(request, "Both start and end dates are required.", extra_tags="semester")
            return redirect("class_management")

        try:
            start_date = datetime.strptime(start, "%Y-%m-%d").date()
            end_date = datetime.strptime(end, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid date format.", extra_tags="semester")
            return redirect("class_management")

        if end_date <= start_date:
            messages.error(request, "End date must be after start date.", extra_tags="semester")
            return redirect("class_management")

        if start_date < today:
            messages.error(request, "Semester start date cannot be earlier than today.", extra_tags="semester")
            return redirect("class_management")

        # If a semester exists and is ongoing → block unless editing
        if current_semester and "confirm_edit" not in request.POST:
            messages.error(request, "A semester is already active. Confirm edit to change it.", extra_tags="semester")
            return redirect("class_management")

        # If editing, update existing; else create new
        if current_semester and "confirm_edit" in request.POST:
            current_semester.start_date = start_date
            current_semester.end_date = end_date
            current_semester.save()
            messages.success(request, "Semester period updated successfully.", extra_tags="semester")
        else:
            Semester.objects.all().delete()  # make sure only one exists
            current_semester = Semester.objects.create(start_date=start_date, end_date=end_date)
            messages.success(request, "Semester period saved successfully.", extra_tags="semester")

        return redirect("class_management")

    # this stays at the bottom so the page renders when not POST
    return render(request, "class_management.html", {"current_semester": current_semester,"today": today })

# API views for mobile app integration
class AccountViewSet(viewsets.ModelViewSet):
    queryset = Account.objects.all()
    serializer_class = AccountSerializer
    
    def get_permissions(self):
        """Require authentication for write operations"""
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsAuthenticated()]
    
    def create(self, request, *args, **kwargs):
        """Override create to add debugging"""
        print("=== DEBUG: Incoming account data ===")
        print("Request data:", request.data)
        print("Request headers:", dict(request.headers))
        
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            print("=== DEBUG: Validation errors ===")
            print("Serializer errors:", serializer.errors)
            return Response(serializer.errors, status=400)
        
        return super().create(request, *args, **kwargs) 
    
    def perform_create(self, serializer):
        """Add audit trail and notification for account creation"""
        account = serializer.save()
        
        # Create notification for new account upload
        AccountUploadNotification.objects.create(
            account_name=f"{account.first_name} {account.last_name}"
        )
        
        # Log who created this account
        if self.request.user.username == 'mobile_system':
            print(f"Account {account.user_id} created via mobile upload")
            print(f"Notification created for new account: {account.first_name} {account.last_name}")
        else:
            print(f"Account {account.user_id} created by user {self.request.user.username}")
        
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def mobile_sync(self, request):
        """Sync accounts from mobile to web"""
        serializer = MobileAccountSerializer(data=request.data, many=True)
        if serializer.is_valid():
            return Response({'status': 'success'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Web app to mobile app account overwrite/syncing
@csrf_exempt
@require_http_methods(["GET"])
def mobile_account_sync(request):
    """API endpoint for mobile apps to fetch all accounts"""
    try:
        # Get all accounts from Django database
        accounts = Account.objects.all()
        
        # Convert to list and format for mobile consumption
        accounts_list = []
        for account in accounts:
            accounts_list.append({
                'user_id': account.user_id,
                'email': account.email,
                'first_name': account.first_name,
                'last_name': account.last_name,
                'role': account.role,
                'password': None,  # Don't send passwords
                'sex': account.sex,
                'status': account.status,
                'course_section': account.course_section.id if account.course_section else None,
                'fingerprint_template': account.fingerprint_template
            })
        
        return JsonResponse(accounts_list, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)
        
@csrf_exempt
@require_http_methods(["POST"])
def trigger_mobile_sync(request):
    """Trigger sync to mobile apps"""
    try:
        print("Mobile sync triggered from web admin")
        
        # Get counts of data available for sync
        account_count = Account.objects.count()
        schedule_count = ClassSchedule.objects.count()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Mobile sync triggered successfully',
            'data': {
                'accounts_available': account_count,
                'schedules_available': schedule_count
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

class ClassScheduleViewSet(viewsets.ModelViewSet):
    queryset = ClassSchedule.objects.all()
    serializer_class = MobileClassScheduleSerializer
    
    def get_permissions(self):
        """Allow read without auth, require auth for write operations"""
        if self.action in ['list', 'retrieve', 'today_schedules']:
            return [AllowAny()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'])
    def today_schedules(self, request):
        """Get today's schedules for mobile"""
        today = timezone.now().date()
        schedules = ClassSchedule.objects.all()
        # Make sure this line uses MobileClassScheduleSerializer
        serializer = MobileClassScheduleSerializer(schedules, many=True)
        return Response(serializer.data)

class AttendanceRecordViewSet(viewsets.ModelViewSet):
    queryset = AttendanceRecord.objects.all()
    serializer_class = AttendanceRecordSerializer
    
    def get_permissions(self):
        """Allow read without auth, require auth for write operations"""
        if self.action in ['list', 'retrieve', 'download_for_mobile']:
            return [AllowAny()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def mobile_upload(self, request):
        """Upload attendance records from mobile"""
        # Explicitly use MobileAttendanceSerializer (which has the user_id lookup logic)
        serializer = MobileAttendanceSerializer(data=request.data, many=True)
        if serializer.is_valid():
            created_records = serializer.save()
            return Response({
                'status': 'success', 
                'count': len(created_records),
                'message': f'Successfully uploaded {len(created_records)} attendance records'
            })
        else:
            print("Validation errors:", serializer.errors)  # Debug print
            return Response({
                'status': 'error',
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def download_for_mobile(self, request):
        """Download attendance records to mobile"""
        date_param = request.query_params.get('date')
        if date_param:
            try:
                filter_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                records = AttendanceRecord.objects.filter(date=filter_date)
            except ValueError:
                return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            records = AttendanceRecord.objects.all()
        
        serializer = MobileAttendanceSerializer(records, many=True)
        return Response(serializer.data)

@api_view(['POST'])
@permission_classes([AllowAny])
def mobile_login(request):
    """Login endpoint for mobile app"""
    username = request.data.get('username')
    password = request.data.get('password')
    
    if username and password:
        user = authenticate(username=username, password=password)
        if user:
            token, created = Token.objects.get_or_create(user=user)
            return Response({
                'token': token.key,
                'user_id': user.id,
                'username': user.username
            })
    
    return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

# For docx file generation
@instructor_or_admin_required
def generate_attendance_docx_view(request, class_id):
    """
    Generate attendance template as .docx file
    """
    try:
        # Path to your template file
        template_path = os.path.join(settings.BASE_DIR, 'PTLT_App', 'templates', 'attendance_template.docx')
        
        # Load the template
        doc = DocxTemplate(template_path)
        
        # Get class schedule data
        class_schedule = get_object_or_404(ClassSchedule, id=class_id)
        
        # Get students for this class
        students = Account.objects.filter(
            course_section=class_schedule.course_section, 
            role='Student'
        ).order_by('last_name', 'first_name')
        
        # Generate dates
        dates = []
        base_date = date.today()
        for i in range(8):
            new_date = base_date + timedelta(days=i*2)
            dates.append(new_date.strftime("%m/%d"))
        
        context = {
            'subject': class_schedule.course_title,
            'faculty_name': f"{class_schedule.professor.first_name} {class_schedule.professor.last_name}" if class_schedule.professor else '',
            'course': class_schedule.course_section.course_name if class_schedule.course_section else '',
            'room_assignment': class_schedule.room_assignment,
            'year_section': class_schedule.course_section.course_section if class_schedule.course_section else '',
            'schedule': f"{class_schedule.days} {class_schedule.time_in}-{class_schedule.time_out}",
            'date1': dates[0],
            'date2': dates[1],
            'date3': dates[2],
            'date4': dates[3],
            'date5': dates[4],
            'date6': dates[5],
            'date7': dates[6],
            'date8': dates[7],
        }

        # Add individual student data for up to 40 students
        for i, student in enumerate(students[:40]):  # Limit to first 40 students
            context[f'student{i+1}_name'] = f"{student.last_name}, {student.first_name}"
            context[f'student{i+1}_sex'] = student.sex
            
            # Add empty time slots for now
            for j in range(1, 9):
                context[f'student{i+1}_time{j}'] = ''

        # Fill remaining slots with empty strings if fewer than 40 students
        for i in range(len(students), 40):
            context[f'student{i+1}_name'] = ''
            context[f'student{i+1}_sex'] = ''
            for j in range(1, 9):
                context[f'student{i+1}_time{j}'] = ''                    
        
        # Render the template
        doc.render(context)
        
        # Save to memory
        file_stream = BytesIO()
        doc.save(file_stream)
        file_stream.seek(0)
        
        # Create response
        response = HttpResponse(
            file_stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        
        filename = f"Attendance_{class_schedule.course_code}_{base_date.strftime('%Y%m%d')}.docx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating document: {str(e)}", status=500)
    
@api_view(['POST'])
@permission_classes([AllowAny])
def mobile_auth(request):
    """Authentication endpoint for mobile devices"""
    device_id = request.data.get('device_id')
    device_secret = request.data.get('device_secret')
    
    # You can store these in Django settings or environment variables
    VALID_DEVICES = {
        'FINGERPRINT_DEVICE_001': 'Room1_Debug_2025', 
        # Add more devices as needed
    }
    
    if device_id in VALID_DEVICES and VALID_DEVICES[device_id] == device_secret:
        # Create or get a system user for mobile uploads
        user, created = User.objects.get_or_create(
            username='mobile_system',
            defaults={'email': 'mobile@system.local', 'is_active': True}
        )
        
        # Create or get token for this user
        token, created = Token.objects.get_or_create(user=user)
        
        return Response({
            'token': token.key,
            'expires_in': 86400,  # 24 hours
            'device_id': device_id
        })
    
    return Response({'error': 'Invalid device credentials'}, status=401)